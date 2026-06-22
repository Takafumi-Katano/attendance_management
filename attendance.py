"""
attendance.py – Excel-based attendance record management.

Handles:
- Creating / opening Attendance_Sheet_YYYY.xlsx files.
- Adding start / end times to the correct year-file and month-sheet.
- Calculating work time (with 1-hour break deduction when > 6 h).
- Sorting rows by date ascending.
- Writing a monthly total row when all days in the month are present.
"""

import calendar
import glob
import os
from datetime import date, datetime, time
from typing import List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from app_logger import get_logger

HEADERS = ["日付", "始業時間", "終業時間", "労働時間"]
DATE_FMT = "%Y/%m/%d"
TIME_FMT = "%H:%M"
logger = get_logger()


class AttendanceManager:
    """Reads and writes attendance data to Excel files."""

    def __init__(self, config, task_session_manager=None) -> None:
        self.config = config
        self.task_session_manager = task_session_manager

    # ------------------------------------------------------------------
    # Path helpers
    # ------------------------------------------------------------------

    def get_file_path(self, year: Optional[int] = None) -> Optional[str]:
        """Return the full path of the Excel file for *year* (default: current year)."""
        if not self.config.folder_path:
            return None
        if year is None:
            year = datetime.now().year
        return os.path.join(
            self.config.folder_path, f"Attendance_Sheet_{year}.xlsx"
        )

    @staticmethod
    def get_sheet_name(month: Optional[int] = None) -> str:
        """Return the abbreviated month name used as the sheet name (e.g. 'Jan')."""
        if month is None:
            month = datetime.now().month
        return datetime(2000, month, 1).strftime("%b")

    def ensure_file_exists(self, year: Optional[int] = None) -> Optional[str]:
        """Ensure the yearly attendance file exists and return its path."""
        file_path = self.get_file_path(year)
        if not file_path:
            return None
        if os.path.exists(file_path):
            return file_path

        target_year = year if year is not None else datetime.now().year
        now = datetime.now()
        # For non-current years, create Jan sheet as an initial placeholder.
        target_month = now.month if target_year == now.year else 1
        sheet_name = self.get_sheet_name(target_month)

        wb = self._open_or_create_workbook(file_path)
        self._open_or_create_sheet(wb, sheet_name)
        self._set_active_sheet(wb, sheet_name)
        wb.save(file_path)
        wb.close()
        return file_path

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def record_start(self, dt: datetime) -> Optional[str]:
        """Write start time for *dt* to the Excel file.  Returns 'HH:MM' or None."""
        return self._write_time(dt, column=1)

    def record_end(self, dt: datetime) -> Optional[str]:
        """Write end time for *dt* to the Excel file.  Returns 'HH:MM' or None."""
        return self._write_time(dt, column=2)

    def get_today_times(self) -> Tuple[Optional[str], Optional[str]]:
        """Return (start_time, end_time) strings for today, or (None, None)."""
        now = datetime.now()
        file_path = self.get_file_path(now.year)
        if not file_path or not os.path.exists(file_path):
            return None, None

        sheet_name = self.get_sheet_name(now.month)
        try:
            wb = load_workbook(file_path, data_only=True)
            try:
                if sheet_name not in wb.sheetnames:
                    return None, None
                ws = wb[sheet_name]
                date_str = now.strftime(DATE_FMT)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and "合計" in str(row[0]):
                        continue
                    row_date = self._normalize_date_cell(row[0])
                    if row_date and row_date == date_str:
                        return self._normalize_time_cell(row[1]), self._normalize_time_cell(row[2])
            finally:
                wb.close()
        except Exception as exc:
            print(f"[AttendanceManager] get_today_times error: {exc}")
            logger.exception("get_today_times error")
        return None, None

    def fill_missing_end_time(self, target_date: date, end_dt: datetime) -> bool:
        """Fill a missing end time for *target_date* (date object) with *end_dt*.

        Returns True if the record was updated, False otherwise.
        """
        file_path = self.get_file_path(target_date.year)
        if not file_path or not os.path.exists(file_path):
            logger.info(
                "fill_missing_end_time skipped: file not found for %s (%s)",
                target_date.isoformat(),
                file_path,
            )
            return False

        sheet_name = self.get_sheet_name(target_date.month)
        try:
            wb = load_workbook(file_path)
            try:
                if sheet_name not in wb.sheetnames:
                    logger.info(
                        "fill_missing_end_time skipped: sheet '%s' not found in %s",
                        sheet_name,
                        file_path,
                    )
                    return False
                ws = wb[sheet_name]
                data = self._read_data(ws)
                date_str = target_date.strftime(DATE_FMT)
                for row in data:
                    if row[0] == date_str and row[1] and not row[2]:
                        end_str = end_dt.strftime(TIME_FMT)
                        logger.info("fill_missing_end_time updating: date=%s end=%s", date_str, end_str)
                        row[2] = end_str
                        self._flush(ws, data, target_date.year, target_date.month)
                        logger.info("fill_missing_end_time recalculated work-time using sheet formulas")
                        self._set_active_sheet(wb, sheet_name)
                        wb.save(file_path)
                        logger.info(
                            "fill_missing_end_time success: %s end=%s file=%s",
                            date_str,
                            end_str,
                            file_path,
                        )
                        return True
                logger.info(
                    "fill_missing_end_time skipped: missing target row for %s",
                    target_date.isoformat(),
                )
            finally:
                wb.close()
        except Exception as exc:
            print(f"[AttendanceManager] fill_missing_end_time error: {exc}")
            logger.exception("fill_missing_end_time error for %s", target_date.isoformat())
        return False

    def find_latest_missing_end_date(self) -> Optional[date]:
        """Scan all Attendance_Sheet_*.xlsx files and return the latest date that has
        a start time but no end time.  Returns None if no such date exists or the
        folder is not configured.
        """
        if not self.config.folder_path:
            logger.info("find_latest_missing_end_date skipped: folder path is not configured")
            return None

        pattern = os.path.join(self.config.folder_path, "Attendance_Sheet_*.xlsx")
        today = datetime.now().date()
        latest: Optional[date] = None
        logger.info("find_latest_missing_end_date scanning: pattern=%s", pattern)

        for file_path in glob.glob(pattern):
            try:
                wb = load_workbook(file_path, data_only=True)
                try:
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row[0] and "合計" in str(row[0]):
                                continue
                            date_val = self._normalize_date_cell(row[0])
                            start_val = self._normalize_time_cell(row[1])
                            end_val = self._normalize_time_cell(row[2])
                            if date_val and start_val and not end_val:
                                d = datetime.strptime(date_val, DATE_FMT).date()
                                if d < today and (latest is None or d > latest):
                                    latest = d
                finally:
                    wb.close()
            except Exception as exc:
                print(f"[AttendanceManager] find_latest_missing_end_date error reading {file_path}: {exc}")
                logger.exception(
                    "find_latest_missing_end_date error reading file: %s",
                    file_path,
                )

        logger.info("find_latest_missing_end_date result: %s", latest.isoformat() if latest else "None")
        return latest

    def check_previous_day(self) -> None:
        """Find the most recent workday with a missing end time and auto-fill it
        from Windows Event Log.

        Search covers all Attendance_Sheet_*.xlsx files so year/month boundaries
        are handled correctly.
        """
        from windows_events import get_last_work_end_time

        logger.info("check_previous_day started")
        target_date = self.find_latest_missing_end_date()
        if target_date is None:
            logger.info("check_previous_day skipped: no missing end-date found")
            return

        end_time = get_last_work_end_time(target_date)
        if end_time is None:
            logger.info(
                "check_previous_day skipped: no candidate event time found for %s",
                target_date.isoformat(),
            )
            return

        # Requirement update: treat qualifying lock-time events as end-of-work even
        # when the lock spans overnight, so only the date must match here.
        if end_time.date() != target_date:
            logger.info(
                "check_previous_day skipped: candidate out of date (%s not on %s)",
                end_time.strftime("%Y-%m-%d %H:%M:%S"),
                target_date.isoformat(),
            )
            return

        try:
            updated = self.fill_missing_end_time(target_date, end_time)
            logger.info(
                "check_previous_day finished: date=%s candidate=%s updated=%s",
                target_date.isoformat(),
                end_time.strftime("%Y-%m-%d %H:%M:%S"),
                updated,
            )
        except Exception as exc:
            print(f"[AttendanceManager] check_previous_day error: {exc}")
            logger.exception("check_previous_day error for %s", target_date.isoformat())

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _write_time(self, dt: datetime, column: int) -> Optional[str]:
        """Internal: write start (column=1) or end (column=2) time."""
        if not self.config.folder_path:
            logger.info("_write_time skipped: folder path not configured")
            return None

        file_path = self.get_file_path(dt.year)
        sheet_name = self.get_sheet_name(dt.month)

        wb = self._open_or_create_workbook(file_path)
        ws = self._open_or_create_sheet(wb, sheet_name)

        date_str = dt.strftime(DATE_FMT)
        time_str = dt.strftime(TIME_FMT)

        data = self._read_data(ws)

        # Find existing row for today or create a new one
        target_row = None
        for row in data:
            if row[0] == date_str:
                target_row = row
                break
        if target_row is None:
            target_row = [date_str, None, None, None]
            data.append(target_row)
            logger.info("_write_time created new row: date=%s", date_str)

        target_row[column] = time_str
        logger.info(
            "_write_time updated: date=%s column=%s value=%s",
            date_str,
            "start" if column == 1 else "end",
            time_str,
        )

        self._flush(ws, data, dt.year, dt.month)
        self._set_active_sheet(wb, sheet_name)
        wb.save(file_path)
        wb.close()
        logger.info("_write_time saved: file=%s sheet=%s", file_path, sheet_name)
        return time_str

    @staticmethod
    def _open_or_create_workbook(file_path: str) -> Workbook:
        if os.path.exists(file_path):
            return load_workbook(file_path)
        wb = Workbook()
        wb.remove(wb.active)  # Remove the default empty sheet
        return wb

    @staticmethod
    def _open_or_create_sheet(wb: Workbook, sheet_name: str):
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        ws.append(HEADERS)
        # Basic column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws["A1"].number_format = "yyyy/mm/dd"
        ws["B1"].number_format = "hh:mm"
        ws["C1"].number_format = "hh:mm"
        ws["D1"].number_format = "[h]:mm"
        return ws

    @staticmethod
    def _set_active_sheet(wb: Workbook, sheet_name: str) -> None:
        if sheet_name in wb.sheetnames:
            wb.active = wb.sheetnames.index(sheet_name)
        else:
            print(f"[AttendanceManager] _set_active_sheet warning: sheet not found: {sheet_name}")

    def _read_data(self, ws) -> list:
        """Read data rows (skip header and total rows) as a list of lists."""
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and "合計" not in str(row[0]):
                date_str = self._normalize_date_cell(row[0])
                if not date_str:
                    continue
                rows.append(
                    [
                        date_str,
                        self._normalize_time_cell(row[1]),
                        self._normalize_time_cell(row[2]),
                        None,
                    ]
                )
        return rows

    @staticmethod
    def _normalize_date_cell(value) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.strftime(DATE_FMT)
        if isinstance(value, date):
            return value.strftime(DATE_FMT)
        text = str(value).strip()
        if not text:
            return None
        for fmt in (DATE_FMT, "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).strftime(DATE_FMT)
            except ValueError:
                continue
        return None

    @staticmethod
    def _normalize_time_cell(value) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.strftime(TIME_FMT)
        if isinstance(value, time):
            return value.strftime(TIME_FMT)
        text = str(value).strip()
        if not text:
            return None
        for fmt in (TIME_FMT, "%H:%M:%S"):
            try:
                return datetime.strptime(text, fmt).strftime(TIME_FMT)
            except ValueError:
                continue
        return None

    @staticmethod
    def _work_time_formula(row_num: int) -> str:
        # Handle both numeric time cells and text time cells, including overnight
        # spans via MOD(), then apply a 1-hour break deduction when over 6 hours.
        b = f"B{row_num}"
        c = f"C{row_num}"
        diff = (
            f"(IF(ISNUMBER({c}),{c},TIMEVALUE({c}))"
            f"-IF(ISNUMBER({b}),{b},TIMEVALUE({b})))"
        )
        return (
            f'=IF(OR({b}="",{c}=""),"",'
            f"IFERROR(MOD({diff},1)-IF(MOD({diff},1)*24>6,1/24,0),\"\"))"
        )

    def _flush(self, ws, data: list, year: int, month: int) -> None:
        """Write sorted data back to *ws*, adding a total row if month is full."""
        # Sort by date string (YYYY/MM/DD sorts lexicographically)
        data.sort(key=lambda r: str(r[0]) if r[0] else "")

        # Clear all existing data rows (keep header in row 1)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        # Write sorted data (cols A–D)
        for idx, row in enumerate(data, start=2):
            ws.cell(row=idx, column=1, value=row[0])
            ws.cell(row=idx, column=2, value=row[1])
            ws.cell(row=idx, column=3, value=row[2])
            work_time_cell = ws.cell(row=idx, column=4, value=self._work_time_formula(idx))
            work_time_cell.number_format = "[h]:mm"
            ws.cell(row=idx, column=1).number_format = "yyyy/mm/dd"
            ws.cell(row=idx, column=2).number_format = "hh:mm"
            ws.cell(row=idx, column=3).number_format = "hh:mm"

        # ── Task columns (E+) ─────────────────────────────────────────────
        # Clear existing task column headers from row 1 (cols 5+)
        col_idx = 5
        while ws.cell(row=1, column=col_idx).value is not None:
            ws.cell(row=1, column=col_idx).value = None
            col_idx += 1

        tasks_for_month: List[str] = []
        if self.task_session_manager is not None:
            tasks_for_month = self.task_session_manager.get_first_occurrence_for_month(
                year, month
            )

        if tasks_for_month:
            task_col_map = {task: 5 + i for i, task in enumerate(tasks_for_month)}

            # Write task header row (row 1, cols E+)
            for task, col in task_col_map.items():
                col_letter = get_column_letter(col)
                ws.cell(row=1, column=col, value=task)
                ws.column_dimensions[col_letter].width = 12

            # Write task duration data for each date row
            for data_idx, row in enumerate(data, start=2):
                date_str = row[0]
                if not date_str:
                    continue
                durations = self.task_session_manager.get_task_durations_for_date(date_str)
                for task, col in task_col_map.items():
                    dur = durations.get(task, 0.0)
                    if dur > 0.0:
                        cell = ws.cell(row=data_idx, column=col, value=dur)
                        cell.number_format = "[h]:mm"

        # ── Total row ────────────────────────────────────────────────────
        last_day = calendar.monthrange(year, month)[1]
        if len(data) == last_day:
            total_row = len(data) + 2
            ws.cell(row=total_row, column=1, value="合計")
            ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row - 1})")
            ws.cell(row=total_row, column=4).number_format = "[h]:mm"
            if tasks_for_month:
                for i, task in enumerate(tasks_for_month):
                    col = 5 + i
                    col_letter = get_column_letter(col)
                    ws.cell(
                        row=total_row,
                        column=col,
                        value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})",
                    )
                    ws.cell(row=total_row, column=col).number_format = "[h]:mm"
